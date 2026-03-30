#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import asyncio
import openpyxl
import aiohttp
from io import BytesIO
from pyrogram import Client, filters
from pyrogram.types import Message, InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery


class CashBackBot:
    """Бот для кэшбэка с чтением Excel по HTTP"""
    
    # Целевая зеленая заливка (RGB: 146,208,80 -> HEX: 92D050)
    TARGET_COLOR_RGB = "92D050"
    
    def __init__(self):
        """Загрузка конфигурации: приоритет bot.conf > переменные окружения"""
        print("🚀 Инициализация CashBackBot...")
        
        # Инициализируем переменные
        self.token = None
        self.api_id = None
        self.api_hash = None
        self.file_url = None
        self.file_user = None
        self.file_pass = None
        self.categories = []
        self.category_emojis_dict = {}
        self.cards = []
        self.row_data = {}
        self.comments_data = {}  # Словарь для примечаний к процентам
        self.current_page = 0
        
        # Пытаемся загрузить из bot.conf
        config_loaded = self._load_config_from_file('bot.conf')
        
        if config_loaded:
            print("📁 Конфигурация загружена из bot.conf")
        else:
            print("📁 Файл bot.conf не найден, используем переменные окружения")
            self._load_config_from_env()
        
        # Директория для сессий
        self.sessions_dir = os.environ.get('SESSIONS_DIR', '/app/sessions')
        os.makedirs(self.sessions_dir, exist_ok=True)
        
        # Проверка обязательных переменных
        if not self.token:
            raise ValueError("❌ BOT_TOKEN not found")
        
        if not self.api_id or not self.api_hash:
            raise ValueError("❌ API_ID and API_HASH are required! Get them from https://my.telegram.org/apps")
        
        if not self.file_url:
            print("⚠️ EXCEL_URL not set, bot will work without data")
        
        # Создаем клиента
        print("🔌 Создание клиента Telegram...")
        self.app = Client(
            name=f"{self.sessions_dir}/cashback_bot_session",
            api_id=self.api_id,
            api_hash=self.api_hash,
            bot_token=self.token,
            workdir="."
        )
        
        print("✅ Инициализация завершена")
    
    def _load_config_from_file(self, config_file):
        """Загрузка конфигурации из файла"""
        if not os.path.exists(config_file):
            return False
        
        print(f"📁 Чтение конфигурации из {config_file}...")
        try:
            with open(config_file, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith('#'):
                        continue
                    
                    if '=' in line:
                        key, value = line.split('=', 1)
                        key = key.strip().lower()
                        value = value.strip()
                        
                        if key == 'token':
                            self.token = value
                            print("   ✅ Токен загружен из файла")
                        elif key == 'api_id':
                            self.api_id = int(value)
                            print("   ✅ API ID загружен из файла")
                        elif key == 'api_hash':
                            self.api_hash = value
                            print("   ✅ API Hash загружен из файла")
                        elif key == 'file_url':
                            self.file_url = value
                            print(f"   ✅ URL файла: {self.file_url}")
                        elif key == 'file_user':
                            self.file_user = value
                            print("   ✅ Логин загружен")
                        elif key == 'file_pass':
                            self.file_pass = value
                            print("   ✅ Пароль загружен")
            return True
        except Exception as e:
            print(f"   ⚠️ Ошибка чтения конфига: {e}")
            return False
    
    def _load_config_from_env(self):
        """Загрузка конфигурации из переменных окружения"""
        self.token = os.environ.get('BOT_TOKEN')
        self.api_id = os.environ.get('API_ID')
        self.api_hash = os.environ.get('API_HASH')
        self.file_url = os.environ.get('EXCEL_URL')
        self.file_user = os.environ.get('EXCEL_USER')
        self.file_pass = os.environ.get('EXCEL_PASS')
        
        if self.token:
            print("   ✅ Токен загружен из переменных окружения")
        if self.api_id:
            print("   ✅ API ID загружен из переменных окружения")
        if self.api_hash:
            print("   ✅ API Hash загружен из переменных окружения")
        if self.file_url:
            print(f"   ✅ URL файла: {self.file_url}")
        if self.file_user:
            print("   ✅ Логин загружен из переменных окружения")
        if self.file_pass:
            print("   ✅ Пароль загружен из переменных окружения")
    
    async def _download_excel(self):
        """Скачивание Excel файла по HTTP с авторизацией"""
        if not self.file_url:
            print("⚠️ URL файла не указан")
            return None
        
        try:
            print(f"📥 Скачивание файла: {self.file_url}")
            
            auth = None
            if self.file_user and self.file_pass:
                auth = aiohttp.BasicAuth(self.file_user, self.file_pass)
            
            async with aiohttp.ClientSession() as session:
                async with session.get(self.file_url, auth=auth, timeout=30) as response:
                    if response.status == 200:
                        data = await response.read()
                        print(f"✅ Файл скачан, размер: {len(data)} байт")
                        return BytesIO(data)
                    elif response.status == 401:
                        print("❌ Ошибка авторизации: 401 Unauthorized")
                        return None
                    elif response.status == 404:
                        print("❌ Файл не найден: 404 Not Found")
                        return None
                    else:
                        print(f"❌ Ошибка скачивания: HTTP {response.status}")
                        return None
                        
        except asyncio.TimeoutError:
            print("❌ Таймаут при скачивании файла")
            return None
        except Exception as e:
            print(f"❌ Ошибка при скачивании: {e}")
            return None
    
    def _get_cell_color(self, cell):
        """Получить цвет заливки ячейки"""
        fill = cell.fill
        if fill and fill.fgColor:
            color = fill.fgColor
            if color.rgb:
                return color.rgb
        return None
    
    def _is_green_color(self, color):
        """Проверяет, является ли цвет целевым зеленым (RGB: 146,208,80)"""
        if not color:
            return False
        
        color_upper = color.upper()
        
        # Проверяем различные форматы
        if "92D050" in color_upper:
            return True
        if "FF92D050" in color_upper:
            return True
        
        return False
    
    def _get_cell_comment(self, cell):
        """Получить текст примечания из ячейки"""
        if cell.comment:
            return cell.comment.text.strip()
        return None
    
    def _parse_excel(self, file_io):
        """Парсинг Excel файла (лист 'ИсходныеДанные') с чтением примечаний"""
        try:
            print("📖 Парсинг Excel файла...")
            workbook = openpyxl.load_workbook(file_io, data_only=True)
            
            if "ИсходныеДанные" not in workbook.sheetnames:
                print(f"⚠️ Лист 'ИсходныеДанные' не найден. Доступные листы: {workbook.sheetnames}")
                sheet = workbook.worksheets[0]
                print(f"📋 Используется первый лист: {sheet.title}")
            else:
                sheet = workbook["ИсходныеДанные"]
                print(f"📋 Используется лист: ИсходныеДанные")
            
            # Читаем ячейку B1 - количество строк
            loop_count = sheet['B1'].value
            if loop_count is None:
                loop_count = 0
            else:
                loop_count = int(loop_count)
            
            print(f"📊 Счетчик строк (B1): {loop_count}")
            
            # Читаем названия карт
            self.cards = []
            col_idx = 3
            while col_idx <= 100:
                card_name = sheet.cell(row=1, column=col_idx).value
                if card_name:
                    self.cards.append(str(card_name).strip())
                col_idx += 1
            
            print(f"💳 Найдено карт: {len(self.cards)}")
            
            # Основной парсинг
            self.categories = []
            self.category_emojis_dict = {}
            self.row_data = {}
            self.comments_data = {}
            
            for row_num in range(2, loop_count + 2):
                emoji = sheet.cell(row=row_num, column=1).value
                category = sheet.cell(row=row_num, column=2).value
                
                if not category:
                    continue
                
                category = str(category).strip()
                emoji_str = str(emoji).strip() if emoji else ''
                
                has_valid = False
                values = {}
                comments = {}
                
                for card_idx, card_name in enumerate(self.cards):
                    col_idx = 3 + card_idx
                    cell = sheet.cell(row=row_num, column=col_idx)
                    value = cell.value
                    color = self._get_cell_color(cell)
                    is_green = self._is_green_color(color)
                    
                    # Читаем примечание к ячейке с процентом
                    comment = self._get_cell_comment(cell)
                    
                    if value is not None:
                        try:
                            if isinstance(value, str):
                                value = value.replace('%', '').strip()
                            percent = float(value)
                        except (ValueError, TypeError):
                            percent = 0
                    else:
                        percent = 0
                    
                    if is_green and percent > 0:
                        values[card_name] = percent
                        if comment:
                            comments[card_name] = comment
                        has_valid = True
                
                if has_valid:
                    self.categories.append(category)
                    self.category_emojis_dict[category] = emoji_str
                    self.row_data[category] = values
                    self.comments_data[category] = comments
            
            workbook.close()
            
            # Сортировка: "Все покупки" на первом месте, остальные по алфавиту
            def sort_key(cat):
                cat_stripped = cat.strip().lower()
                if cat_stripped == "все покупки":
                    return (0, "")
                return (1, cat_stripped)
            
            self.categories.sort(key=sort_key)
            
            print(f"✅ Загружено {len(self.categories)} категорий")
            return True
            
        except Exception as e:
            print(f"❌ Ошибка парсинга Excel: {e}")
            return False
    
    async def _load_data(self):
        """Загрузка и парсинг Excel файла"""
        print("📥 Загрузка данных...")
        file_io = await self._download_excel()
        if file_io:
            return self._parse_excel(file_io)
        return False
    
    def get_categories_keyboard(self, page=0, items_per_page=10):
        """Создание клавиатуры с категориями"""
        if not self.categories:
            return None
        
        start_idx = page * items_per_page
        end_idx = start_idx + items_per_page
        page_categories = self.categories[start_idx:end_idx]
        
        buttons = []
        for cat in page_categories:
            emoji = self.category_emojis_dict.get(cat, '')
            button_text = f"{emoji} {cat}" if emoji else cat
            buttons.append([InlineKeyboardButton(button_text, callback_data=f"cat_{cat}")])
        
        nav_buttons = []
        total_pages = (len(self.categories) + items_per_page - 1) // items_per_page
        
        if page > 0:
            nav_buttons.append(InlineKeyboardButton("◀️ Назад", callback_data=f"page_{page-1}"))
        if page < total_pages - 1:
            nav_buttons.append(InlineKeyboardButton("Вперед ▶️", callback_data=f"page_{page+1}"))
        if nav_buttons:
            buttons.append(nav_buttons)
        
        buttons.append([InlineKeyboardButton("❌ Закрыть", callback_data="close")])
        
        return InlineKeyboardMarkup(buttons)
    
    def get_category_info(self, category_name):
        """Получить отсортированную информацию по категории с примечаниями"""
        if category_name not in self.row_data:
            return None
        
        values = self.row_data[category_name]
        comments = self.comments_data.get(category_name, {})
        
        # Собираем данные с примечаниями
        items = []
        for card, value in values.items():
            comment = comments.get(card, '')
            items.append((card, value, comment))
        
        # Сортируем по убыванию процента
        items.sort(key=lambda x: x[1], reverse=True)
        
        return items
    
    async def start(self):
        """Запуск бота"""
        try:
            print("🚀 Запуск CashBackBot...")
            print("🔌 Подключение к Telegram...")
            await self.app.start()
            
            me = await self.app.get_me()
            print(f"✅ Бот запущен: @{me.username if me.username else me.first_name}")
            print(f"   ID: {me.id}")
            print(f"   Имя: {me.first_name}")
            
            print("🤖 CashBackBot готов к работе!")
            print("📝 Нажмите Ctrl+C для остановки")
            print("=" * 50)
            
            self._register_handlers()
            await asyncio.Event().wait()
            
        except Exception as e:
            print(f"❌ Ошибка при запуске: {e}")
            raise
    
    def _register_handlers(self):
        """Регистрация обработчиков сообщений"""
        
        @self.app.on_message(filters.command("start"))
        async def start_command(client, message: Message):
            print(f"📩 Получена команда /start от {message.from_user.id}")
            
            status_msg = await message.reply("🔄 Загрузка данных...")
            
            success = await self._load_data()
            
            if not success:
                await status_msg.edit("❌ Ошибка загрузки данных")
                return
            
            if not self.categories:
                await status_msg.edit("❌ Нет категорий с зелеными процентами")
                return
            
            keyboard = self.get_categories_keyboard(page=0)
            await status_msg.delete()
            await message.reply(
                "💰 **Добро пожаловать в CashBackBot!**\n\n📋 **Выберите категорию:**",
                reply_markup=keyboard
            )
            print(f"✅ Отправлен список категорий ({len(self.categories)} шт)")
        
        @self.app.on_callback_query()
        async def handle_callback(client, callback_query: CallbackQuery):
            data = callback_query.data
            
            if data.startswith("cat_"):
                category_name = data[4:]
                print(f"🔍 Выбрана категория: {category_name}")
                
                items = self.get_category_info(category_name)
                
                if not items:
                    await callback_query.answer("Нет данных по категории", show_alert=True)
                    return
                
                emoji = self.category_emojis_dict.get(category_name, '')
                response = f"{emoji} **{category_name}**\n\n" if emoji else f"**{category_name}**\n\n"
                response += "**💰 Кэшбэк по картам:**\n\n"
                
                for card, value, comment in items:
                    line = f"• **{card}**: {value}%"
                    if comment:
                        line += f" ({comment})"
                    response += line + "\n"
                
                await callback_query.answer()
                await callback_query.message.edit(
                    response,
                    reply_markup=InlineKeyboardMarkup([[
                        InlineKeyboardButton("◀️ Назад к категориям", callback_data="back_to_categories")
                    ]])
                )
                print(f"✅ Отправлена информация по категории: {category_name}")
            
            elif data.startswith("page_"):
                page = int(data[5:])
                self.current_page = page
                print(f"📄 Переход на страницу {page + 1}")
                
                keyboard = self.get_categories_keyboard(page=page)
                if keyboard:
                    await callback_query.answer()
                    await callback_query.message.edit(
                        "💰 **Добро пожаловать в CashBackBot!**\n\n📋 **Выберите категорию:**",
                        reply_markup=keyboard
                    )
            
            elif data == "back_to_categories":
                print("🔙 Возврат к списку категорий")
                keyboard = self.get_categories_keyboard(page=self.current_page)
                if keyboard:
                    await callback_query.answer()
                    await callback_query.message.edit(
                        "💰 **Добро пожаловать в CashBackBot!**\n\n📋 **Выберите категорию:**",
                        reply_markup=keyboard
                    )
            
            elif data == "close":
                print("❌ Закрытие сообщения")
                await callback_query.answer()
                await callback_query.message.delete()
    
    async def stop(self):
        """Остановка бота"""
        if hasattr(self, 'app') and self.app:
            print("\n🛑 Останавливаем CashBackBot...")
            await self.app.stop()
            print("✅ CashBackBot остановлен")


async def main():
    """Основная функция"""
    bot = None
    try:
        bot = CashBackBot()
        await bot.start()
    except KeyboardInterrupt:
        print("\n⚠️ Получен сигнал остановки (Ctrl+C)")
    except Exception as e:
        print(f"❌ Критическая ошибка: {e}")
        import traceback
        traceback.print_exc()
        return 1
    finally:
        if bot:
            await bot.stop()
    
    return 0


if __name__ == '__main__':
    if sys.platform == 'win32':
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    
    asyncio.run(main())